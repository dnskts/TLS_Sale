# -*- coding: utf-8 -*-
"""Generate docs/format_spec.txt from analyzed Excel data."""
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "docs" / "format_spec.txt"


def col_type(values):
    types = set()
    for value in values:
        if value is None or value == "":
            continue
        if isinstance(value, bool):
            types.add("bool")
        elif isinstance(value, int):
            types.add("int")
        elif isinstance(value, float):
            types.add("float")
        elif isinstance(value, datetime):
            types.add("datetime")
        else:
            types.add("string")
    order = ["datetime", "int", "float", "string", "bool"]
    return sorted(types, key=lambda item: order.index(item) if item in order else 99)


def analyze_workbook(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    header = [str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in rows[0]]
    data_rows = rows[1:]
    row_count = len(data_rows)
    column_values = defaultdict(list)

    for row in data_rows:
        row = list(row) if row else []
        if len(row) < len(header):
            row.extend([None] * (len(header) - len(row)))
        for index, value in enumerate(row[: len(header)]):
            column_values[index].append(value)

    columns = []
    for index, name in enumerate(header):
        values = column_values[index]
        null_count = sum(1 for value in values if value is None or value == "")
        enum = Counter(
            str(value).replace("\xa0", " ").strip()
            for value in values
            if value is not None and value != ""
        )
        samples = set()
        for value in values:
            if value is None or value == "":
                continue
            sample = str(value).replace("\xa0", " ").strip()[:150]
            if len(samples) < 8:
                samples.add(sample)

        columns.append(
            {
                "index": index + 1,
                "name": name if name else f"__col_{index + 1}__",
                "null_pct": round(null_count / row_count * 100, 1) if row_count else 0,
                "types": col_type(values),
                "unique_count": len(enum),
                "top_values": enum.most_common(20),
                "samples": sorted(samples),
            }
        )

    sample_rows = []
    for row in data_rows[:3]:
        row = list(row) if row else []
        if len(row) < len(header):
            row.extend([None] * (len(header) - len(row)))
        sample_rows.append(row[: len(header)])

    return {
        "header": header,
        "row_count": row_count,
        "columns": columns,
        "sample_rows": sample_rows,
    }


def load_data():
    return {
        "1c": analyze_workbook(ROOT / "1C.xlsx", "TDSheet"),
        "bitrix": analyze_workbook(ROOT / "Битрикс.xlsx", "Битрикс"),
    }


def map_type(types):
    if "datetime" in types:
        return "datetime"
    if set(types) <= {"int", "float"}:
        return "number"
    if "int" in types or "float" in types:
        return "number|string"
    return "string"


COL_ALIASES = {
    "Агент": "agent",
    "Поставщик": "supplier",
    "Тип карты": "card_type",
    "Кейс": "case_raw",
    "Канал связи": "channel",
    "Категория": "category",
    "Клиент": "client",
    "Подразделение": "department",
    "Продукт": "product",
    "Наценка": "markup",
    "Заказ": "order_raw",
    "Прибыль": "profit",
    "Прибыль без НДС": "profit_ex_vat",
    "Сервисный сбор": "service_fee",
    "Сумма продажи": "sales_amount",
    "Сумма НДС": "vat_amount",
    "Комиссия": "commission",
    "Комиссия без НДС": "commission_ex_vat",
    "Дата оплаты": "payment_date",
    "Дата реализация": "realization_date",
    "Безнал": "cashless",
    "безнал": "cashless",
    "Сертификат": "certificate",
    "Путешественники": "travelers",
    "Путешественник": "traveler",
    "Номер сделки": "deal_no",
    "Название сделки": "deal_title",
    "Статус сделки": "deal_status",
    "Результат сделки": "deal_result",
    "Тип клиента": "client_type",
    "Менеджер": "manager",
    "Создатель карты": "card_creator",
    "Партнер": "partner",
    "Тип брони": "booking_type",
    "Страна": "country",
    "Город": "city",
    "Гостиница": "hotel",
    "Ресторан": "restaurant",
    "Цепочка": "chain",
    "Количество ночей": "nights_count",
    "Общее количество ночей": "total_nights",
    "Маркетинговый канал": "marketing_channel",
    "Оплата поставщику": "supplier_payment",
    "Удержал поставщик": "supplier_retained",
    "Валюта сделки": "deal_currency",
    "Курс оплаты": "payment_rate",
    "Курс оплаты ЦБ": "cbr_payment_rate",
    "Тип запроса": "request_type",
    "Связанные сделки": "related_deals",
    "Лид": "lead_id",
    "Тур": "tour",
    "Тип оплаты": "payment_type",
    "Депозит": "deposit",
    "Баллы AX": "points_ax",
    "Баллы MR": "points_mr",
    "Баллы IMP": "points_imp",
    "Карта": "card",
    "Код FHR": "fhr_code",
    "Класс": "travel_class",
    "Пассажир": "passenger",
    "Авиакомпания": "airline",
    "Привилегии": "privileges",
    "Наличие договора": "has_contract",
    "Количество сегментов": "segments_count",
    "Кросс-продажа": "cross_sell",
    "Количество номеров": "rooms_count",
    "SR": "sr",
    "LR": "lr",
    "Дополнительная выгода": "additional_benefit",
    "Дополнительная выгода без НДС": "additional_benefit_ex_vat",
    "Сервисный сбор без НДС": "service_fee_ex_vat",
}


def alias_for(name, index):
    if name == "Дата операции":
        return "date_operation" if index == 1 else "datetime_operation"
    if name in COL_ALIASES:
        return COL_ALIASES[name]

    manual = {
        "I d CRM": "id_crm",
        "ID клиента из кейса": "id_client_from_case",
        "ID клиента": "id_client",
        "% участия агента в продаже*": "agent_sale_participation",
        "Номер счёта": "invoice_no",
        "Номер сделки": "deal_no",
        "Схема реализации услуг": "service_scheme",
        "Связанный вид услуги": "related_service_type",
        "Выписывающий агент": "issuing_agent",
        "Ответственное лицо": "responsible_person",
        "Итого оплачено клиентом": "total_paid_by_client",
        "Сумма продажи после возврата": "sales_amount_after_refund",
        "Прибыль РС ТЛС с учетом возврата": "profit_rstls_after_refund",
        "Сумма прибыли с учетом возврата без НДС": "profit_after_refund_ex_vat",
        "Сумма НДС(Комиссии)": "vat_commission",
        "Сумма НДС(Наценка)": "vat_markup",
        "Сумма НДС(Сбор)": "vat_fee",
        "Комиссия поставщика": "supplier_commission",
        "Комиссия поставщика в Валюте": "supplier_commission_currency",
        "Привилегия SOLID BANK": "solid_bank_privilege",
        "Баллы RS Cashback": "rs_cashback_points",
        "В счет ЗП": "against_salary",
        "Убыток на компанию": "loss_company",
        "Убыток на сотрудника": "loss_employee",
        "Кост коды кейса": "case_cost_codes",
        "Связанная компания": "related_company",
        "Дата смены статуса кейса": "case_status_change_date",
        "Клиент из кейса": "client_from_case",
        "Количество броней (1)": "booking_count",
        "Полное наименование организации": "org_full_name",
        "Дата создания фин.карты": "fin_card_created_at",
        "Дата оплаты Клиентом": "client_paid_at",
        "Дата отмены операции (возврат)": "cancel_operation_date",
        "Дата оплаты партнеру (поставщику)": "partner_paid_at",
        "Дата отложенной оплаты": "deferred_payment_date",
        "Сумма отложенной оплаты, руб": "deferred_payment_rub",
        "Сумма отложенной оплаты, валюта": "deferred_payment_currency",
        "Средний курс для возврата": "avg_refund_rate",
        "Сбор поставщика": "supplier_fee",
        "Сбор поставщика на возврат": "supplier_refund_fee",
        "Штраф от поставщика": "supplier_penalty",
        "Штраф клиенту РС ТЛС": "client_penalty_rstls",
        "Возврат сбора РС ТЛС": "fee_refund_rstls",
        "Остаток сбора РС ТЛС": "fee_remainder_rstls",
        "Сумма возврата поставщиком": "supplier_refund_amount",
        "Сумма возврата клиенту": "client_refund_amount",
        "Статус карты возврата": "refund_card_status",
        "Продукты за сбор возврата": "refund_fee_products",
        "Нетто в Валюте поставщика": "net_supplier_currency",
        "Брутто в Валюте поставщика": "gross_supplier_currency",
        "Нетто в рублях": "net_rub",
        "Название валюты сделки": "deal_currency_name",
        "Валюта отложенной оплаты": "deferred_payment_currency_type",
        "Схема финансовой карты": "fin_card_scheme",
        "Кросс-продажа причина": "cross_sell_reason",
        "Комментарий Тимлидеру": "teamlead_comment",
        "Причина стадии Сделка проиграна": "lost_deal_reason",
        "Страна прилета (Конечная точка)": "arrival_country",
        "Город прилета (Конечная точка)": "arrival_city",
        "Дата реализация": "realization_date",
        "Дата оплаты": "payment_date",
        "Дата заезда": "checkin_date",
        "Дата выезда": "checkout_date",
        "Дата вылета": "departure_date",
        "Дата прилета": "arrival_date",
        "Дата начала": "start_date",
        "Дата окончания": "end_date",
        "Дата завершения": "completion_date",
        "Дата оказания услуги": "service_date",
        "Дата создания сделки": "deal_created_at",
        "Дата возврата": "return_date",
        "Дата оплаты Клиентом": "client_paid_at",
        "Дата отмены операции (возврат)": "cancel_operation_date",
        "Дата оплаты партнеру (поставщику)": "partner_paid_at",
        "Дата оплаты Клиентом": "client_paid_at",
    }
    if name in manual:
        return manual[name]

    slug = re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()
    return slug or f"col_{index}"


def field_table(columns):
    lines = [
        "| № | Имя в выгрузке | Alias | Тип | % пустых | Примечание |",
        "|---|----------------|-------|-----|----------|------------|",
    ]
    for col in columns:
        alias = alias_for(col["name"], col["index"])
        note = ""
        if col["null_pct"] >= 95:
            note = "почти всегда пусто"
        elif col["unique_count"] <= 25:
            vals = ", ".join(f"`{v}`" for v, _ in col["top_values"][:10])
            note = f"Значения: {vals}"
        elif col["samples"]:
            note = f"Пример: `{col['samples'][0][:90]}`"
        lines.append(
            f"| {col['index']} | {col['name']} | `{alias}` | {map_type(col['types'])} | "
            f"{col['null_pct']} | {note} |"
        )
    return "\n".join(lines)


def enum_block(columns, name, limit=None):
    col = next(c for c in columns if c["name"] == name)
    lines = [f"#### {name}", ""]
    items = col["top_values"] if limit is None else col["top_values"][:limit]
    for value, count in items:
        lines.append(f"- `{value}` — {count}")
    lines.append("")
    return lines


def sample_block(header, row, count=12):
    lines = []
    for i in range(min(count, len(header))):
        val = row[i]
        if val is None:
            val = ""
        lines.append(f"{header[i]}: {val}")
    return "\n".join(lines)


def main():
    data = load_data()

    c1 = data["1c"]["columns"]
    c2 = data["bitrix"]["columns"]
    h1 = data["1c"]["header"]
    h2 = data["bitrix"]["header"]

    lines = [
        "# Спецификация форматов выгрузок 1С и Битрикс",
        "",
        "Документ описывает структуру входных файлов для парсера и сведения данных в единый дашборд продаж.",
        f"Анализ выполнен по образцам: `1C.xlsx` ({data['1c']['row_count']} строк) и `Битрикс.xlsx` ({data['bitrix']['row_count']} строк).",
        "",
        "---",
        "",
        "## 1. Общие правила загрузки",
        "",
        "### 1.1. Расположение файлов",
        "",
        "- Каталог загрузки: `input/` в корне проекта.",
        "- Ожидаемые имена файлов:",
        "  - `input/1C.xlsx` — выгрузка из 1С",
        "  - `input/Битрикс.xlsx` — выгрузка из Битрикс24 CRM",
        "- Парсер обрабатывает только эти два файла.",
        "",
        "### 1.2. Формат файлов",
        "",
        "| Параметр | Значение |",
        "|----------|----------|",
        "| Формат | Excel Open XML (`.xlsx`) |",
        "| Заголовки | Строка 1 |",
        "| Данные | Со строки 2 |",
        "| Кодировка текста | UTF-8 (кириллица + латиница) |",
        "| Выбор листа | По фиксированному имени, не по индексу |",
        "",
        "### 1.3. Имена листов",
        "",
        "| Файл | Имя листа | Строк данных (образец) | Колонок |",
        "|------|-----------|------------------------|---------|",
        f"| `1C.xlsx` | `TDSheet` | {data['1c']['row_count']} | {len(c1)} |",
        f"| `Битрикс.xlsx` | `Битрикс` | {data['bitrix']['row_count']} | {len(c2)} |",
        "",
        "### 1.4. Гранулярность данных",
        "",
        "| Источник | Единица строки | Комментарий |",
        "|----------|----------------|-------------|",
        "| 1С | Финансовая операция / позиция услуги | Несколько строк на один счёт; до 102 строк на один `0000-XXXXXX` |",
        "| Битрикс | Сделка CRM | Одна строка ≈ одна сделка |",
        "",
        "### 1.5. Период данных в образцах",
        "",
        "- **1С**: 21.03.2025 — 10.07.2026 (по `datetime_operation`)",
        "- **Битрикс**: 04.02.2026 — 10.07.2026 (по `deal_created_at`)",
        "",
        "---",
        "",
        "## 2. Выгрузка 1С (`1C.xlsx`)",
        "",
        "### 2.1. Лист `TDSheet`",
        "",
        "Источник: регистр финансовых операций / продаж в 1С.",
        "",
        "### 2.2. Дублирующийся заголовок `Дата операции`",
        "",
        "| № | Имя в файле | Alias | Формат | Описание |",
        "|---|-------------|-------|--------|----------|",
        "| 1 | Дата операции | `date_operation` | `DD.MM.YYYY` | Дата без времени |",
        "| 2 | Дата операции | `datetime_operation` | `DD.MM.YYYY H:MM:SS` | Дата и время |",
        "",
        "Парсер обязан различать колонки по индексу (1-based), а не по имени.",
        "",
        "### 2.3. Полный перечень полей",
        "",
        field_table(c1),
        "",
        "### 2.4. Ключевые поля для дашборда",
        "",
        "| Поле | Alias | Назначение |",
        "|------|-------|------------|",
        "| Агент | `agent` | Продавец (латиница) |",
        "| Выписывающий агент | `issuing_agent` | Второй участник продажи |",
        "| Клиент | `client` | Плательщик / клиент |",
        "| I d CRM | `id_crm` | ID клиента в CRM |",
        "| Кейс | `case_raw` | Текст кейса Б24 |",
        "| Заказ | `order_raw` | Текст счёта клиенту |",
        "| Подразделение | `department` | Команда агента |",
        "| Связанный вид услуги | `related_service_type` | Тип услуги (RU) |",
        "| Категория | `category` | Категория (EN) |",
        "| Сумма продажи | `sales_amount` | Сумма продажи |",
        "| Прибыль | `profit` | Прибыль |",
        "| Схема реализации услуг | `service_scheme` | SR / LR / агент покупателя |",
        "",
        "### 2.5. Извлекаемые идентификаторы",
        "",
        "| Поле | Regex | Alias результата | Уникальных (образец) |",
        "|------|-------|------------------|----------------------|",
        "| `Кейс` | `000002(\\d+)` | `case_id` | 1 934 |",
        "| `Заказ` | `(0000-\\d+)` | `order_no` | 3 517 |",
        "",
        "Формат `Кейс`: `Кейс 00000210250 от 10.07.2026 16:59:38`",
        "",
        "Формат `Заказ`: `Счет клиенту(Продажа) 0000-057583 от 10.07.2026 19:57:57`",
        "",
        "Связь с Б24: `case_id` = `000002` + `Лид` (zero-pad до 5 цифр). Пример: Лид `10250` → `00000210250`.",
        "",
        "### 2.6. Справочники значений",
        "",
    ]
    lines.extend(enum_block(c1, "Подразделение"))
    lines.extend(enum_block(c1, "Схема реализации услуг"))
    lines.extend(enum_block(c1, "Связанный вид услуги"))
    lines.extend(enum_block(c1, "Категория"))
    lines.extend(enum_block(c1, "Канал связи"))
    lines.extend(enum_block(c1, "Тип карты", 15))

    lines.extend([
        "### 2.7. Правила парсинга 1С",
        "",
        "1. Trim пробелов во всех текстовых полях.",
        "2. Числа: `int`/`float`; пустое = `null`.",
        "3. Даты из строк `DD.MM.YYYY` и `DD.MM.YYYY H:MM:SS`.",
        "4. 665 строк с отрицательной `Сумма продажи` — возвраты; не отбрасывать.",
        "5. 1 762 строки без `Кейс` — операции без CRM-привязки.",
        "6. Не дедуплицировать строки с одним `order_no` без агрегации.",
        "7. `I d CRM` хранить как строку (RSTLS…, числовые ID).",
        "",
        "### 2.8. Примеры строк",
        "",
        "**Пример 1:**",
        "```",
        sample_block(h1, data["1c"]["sample_rows"][0]),
        "```",
        "",
        "**Пример 2:**",
        "```",
        sample_block(h1, data["1c"]["sample_rows"][2]),
        "```",
        "",
        "---",
        "",
        "## 3. Выгрузка Битрикс (`Битрикс.xlsx`)",
        "",
        "### 3.1. Лист `Битрикс`",
        "",
        "Источник: сделки CRM Битрикс24. Одна строка = одна сделка.",
        "",
        "### 3.2. Полный перечень полей",
        "",
        field_table(c2),
        "",
        "### 3.3. Ключевые поля для дашборда",
        "",
        "| Поле | Alias | Назначение |",
        "|------|-------|------------|",
        "| Номер сделки | `deal_no` | PK сделки в CRM |",
        "| Лид | `lead_id` | Основной ключ связи с 1С |",
        "| Ответственное лицо | `responsible_person` | Продавец (кириллица) |",
        "| % участия агента в продаже* | `agent_sale_participation` | Не процент; формат `Имя=сумма` |",
        "| ID клиента | `id_client` | CRM ID клиента |",
        "| Результат сделки | `deal_result` | Успех / Проиграна / В процессе |",
        "| Статус сделки | `deal_status` | Детальный статус воронки |",
        "| Категория | `category` | Категория (RU) |",
        "| Тип запроса | `request_type` | Travel / Lifestyle |",
        "| Сумма продажи | `sales_amount` | Сумма продажи |",
        "| Номер счёта | `invoice_no` | Номер счёта (в образце без `0000-`) |",
        "",
        "### 3.4. Справочники значений",
        "",
    ])
    lines.extend(enum_block(c2, "Результат сделки"))
    lines.extend(enum_block(c2, "Статус сделки"))
    lines.extend(enum_block(c2, "Категория"))
    lines.extend(enum_block(c2, "Тип клиента"))
    lines.extend(enum_block(c2, "Тип запроса"))
    lines.extend(enum_block(c2, "Канал связи"))

    lines.extend([
        "### 3.5. Парсинг `% участия агента в продаже*`",
        "",
        "Поле **не содержит процент**. Формат:",
        "",
        "```",
        "Фамилия Имя=15000.00",
        "```",
        "",
        "Regex: `^(.+?)=(\\d+(?:\\.\\d+)?)$`",
        "",
        "- Группа 1 → `agent_participant_name`",
        "- Группа 2 → `agent_participation_amount`",
        "",
        "Примеры из образца: `Устюжанина Ирина=15000.00`, `Исайкина Ольга=40700.00`.",
        "",
        "### 3.6. Правила парсинга Битрикс",
        "",
        "1. Даты: native Excel `datetime` в колонках 4, 11, 13, 75, 99, 109–111.",
        "2. 1 177 строк с `Сумма продажи = 0` — инфозапросы / без продажи.",
        "3. 737 строк с пустой `Сумма продажи`.",
        "4. Для метрик продаж: фильтр `Результат сделки = 'Успех'` (1 166 строк, ~76.7 млн руб. в образце).",
        "5. Заменять `\xa0` (non-breaking space) на обычный пробел.",
        "6. `ID клиента` может быть int или string.",
        "",
        "### 3.7. Примеры строк",
        "",
        "**Пример 1 — успешная сделка с суммой:**",
        "```",
        sample_block(h2, data["bitrix"]["sample_rows"][min(3, len(data["bitrix"]["sample_rows"]) - 1)]),
        "```",
        "",
        "**Пример 2 — сделка без суммы (инфозапрос):**",
        "```",
        sample_block(h2, data["bitrix"]["sample_rows"][0]),
        "```",
        "",
        "---",
        "",
        "## 4. Сведение 1С ↔ Битрикс",
        "",
        "### 4.1. Ключи связи (приоритет)",
        "",
        "| Приоритет | Поле Б24 | Поле 1С | Алгоритм | Надёжность (образец) |",
        "|-----------|----------|---------|----------|----------------------|",
        "| 1 | `Лид` | `Кейс` | `000002{Лид:05d}` in `Кейс` | 948 / 1 822 (~52%) |",
        "| 2 | `ID клиента` | `I d CRM` | строковое сравнение после trim | 116 общих ID |",
        "| 3 | `Номер счёта` | `Заказ` | regex `0000-\\d+` | 0 совпадений в образце |",
        "| 4 | `Номер сделки` | `Кейс` / `Продукт` | `#38437` в названии сделки | частичное |",
        "",
        "### 4.2. Общие колонки (15 полей с одинаковыми именами)",
        "",
        "`SR`, `LR`, `Баллы AX`, `Баллы IMP`, `Канал связи`, `Категория`, `Клиент`, `Прибыль`, "
        "`Прибыль без НДС`, `Сервисный сбор`, `Сертификат`, `Сумма продажи`, `Тип карты`, "
        "`Убыток на компанию`, `Убыток на сотрудника`.",
        "",
        "Значения могут расходиться: 1С — операция, Б24 — сделка целиком.",
        "",
        "### 4.3. Имена агентов",
        "",
        "| Источник | Формат | Пример |",
        "|----------|--------|--------|",
        "| 1С | Латиница | `Elena Vetvitskaya` |",
        "| Б24 | Кириллица | `Ветвицкая Елена` |",
        "",
        "Парсер хранит оба варианта. Нормализация — отдельная таблица (вне выгрузок).",
        "",
        "### 4.4. Маппинг категорий RU (Б24) ↔ RU/EN (1С)",
        "",
        "| Б24 `Категория` | 1С `Связанный вид услуги` | 1С `Категория` |",
        "|-----------------|---------------------------|----------------|",
        "| Отель | Отельный билет | Accommodation Hotels TR |",
        "| Ресторан | — | Restaurants / Other LS |",
        "| Аренда транспорта c водителем | Трансфер/Авто в распоряжение | Hire&Rental TR / LS |",
        "| Аренда транспорта без водителя | Аренда авто без водителя | Hire&Rental TR |",
        "| Экскурсия | Экскурсия | Sightseeing TR |",
        "| Визы | Виза | — |",
        "| Вип-сервис | ВИП Обслуживание | VIP service at the airport |",
        "| Билеты на мероприятие | Билеты на мероприятие | Tickets |",
        "| — | Авиабилет | Air Tickets |",
        "| — | ЖД билет | Rail Tickets |",
        "",
        "Маппинг частичный; для Travel-услуг предпочтительнее `Связанный вид услуги` из 1С.",
        "",
        "---",
        "",
        "## 5. Качество данных и предупреждения",
        "",
        "- Дублирующийся заголовок `Дата операции` в 1С (колонки 1 и 2).",
        "- Trailing spaces в именах (`Ekaterina Pakulyeva `).",
        "- Non-breaking space `\xa0` в значениях Б24.",
        "- Пустое ≠ ноль в финансовых полях.",
        "- Множественные строки 1С на один счёт — не дедуплицировать.",
        "- Суммы: 1С ~787 млн vs Б24 ~165 млн (разный период и гранулярность).",
        "- `Номер счёта` в Б24 образце не содержит формат `0000-XXXXXX`.",
        "",
        "---",
        "",
        "## 6. Рекомендуемая нормализованная модель парсера",
        "",
        "### 6.1. `operations_1c[]`",
        "",
        "Все строки 1С + вычисляемые поля: `case_id`, `order_no`, `date_operation`, `datetime_operation`.",
        "",
        "### 6.2. `deals_bitrix[]`",
        "",
        "Все сделки Б24 + `lead_id`, `deal_no`, parsed `agent_sale_participation`.",
        "",
        "### 6.3. `links[]`",
        "",
        "```",
        "{ deal_no, lead_id, case_id, id_client, match_type, confidence }",
        "```",
        "",
        "`match_type`: `lead_case` | `client_id` | `invoice` | `deal_text` | `none`",
        "",
        "### 6.4. `agents[]`",
        "",
        "```",
        "{ agent_key, name_1c, name_bitrix, department, group }",
        "```",
        "",
        "### 6.5. Рекомендуемые метрики дашборда",
        "",
        "- Продажи и прибыль по агентам (1С — факт операций; Б24 — успешные сделки).",
        "- Продажи по клиентам (`id_crm` / `id_client`, `client`).",
        "- Продажи по категориям и каналам.",
        "- Сопоставление 1С↔Б24 по `links[]` для сверки.",
        "",
    ])

    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Written {OUT_PATH} ({len(lines)} lines)")


if __name__ == "__main__":
    main()
